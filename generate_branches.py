#!/usr/bin/env python3
"""
Generate 40-branches.mp3 and 40-branches.html from source-data/40-Branches_Chart.xlsx.

Audio structure
───────────────
Before the first slide of each group:
  1 s silence → low gong (220 Hz, long fade) → 300 ms silence
  → [edge-tts male voice] group name → 300 ms
  → for each member in the group:
       [edge-tts] Vedic Name → 500 ms → Quality → 800 ms
  → 1.5 s silence

Per slide:
  4 s silence → 880 Hz bell → 300 ms → "Column: Value" spoken for each non-empty cell
  → brief pause → Vedic Name + Quality repeated once

HTML slideshow stays in sync via JavaScript timeupdate.
Group-header overlay is driven purely by GROUP_HEADER_RANGES timestamps
(no browser speechSynthesis required).

Run:  python generate_branches.py
"""

import asyncio, subprocess, sys, shutil, io, json, re
from pathlib import Path

# ── auto-install Python packages ──────────────────────────────────────────────
def _pip(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", *pkgs])

# audioop was removed in Python 3.13; pydub needs audioop-lts as a replacement
if sys.version_info >= (3, 13):
    try:
        import audioop          # noqa: F401
    except ImportError:
        print("Installing audioop-lts..."); _pip("audioop-lts")

for _pkg, _mod in [
    ("openpyxl",  "openpyxl"),
    ("gtts",      "gtts"),
    ("pydub",     "pydub"),
    ("edge-tts",  "edge_tts"),
]:
    try:
        __import__(_mod)
    except ImportError:
        print(f"Installing {_pkg}..."); _pip(_pkg)

# ── check ffmpeg ───────────────────────────────────────────────────────────────
if not shutil.which("ffmpeg"):
    sys.exit(
        "\nERROR: ffmpeg not found on PATH -- required for MP3 export.\n"
        "  Windows : winget install --id Gyan.FFmpeg\n"
        "  macOS   : brew install ffmpeg\n"
        "  Ubuntu  : sudo apt install ffmpeg\n"
        "After installing, close and reopen this terminal, then run again.\n"
    )

# ── imports ───────────────────────────────────────────────────────────────────
import openpyxl
from gtts import gTTS
from pydub import AudioSegment
from pydub.generators import Sine
import edge_tts

# ── config ────────────────────────────────────────────────────────────────────
XLSX                = Path("source-data/40-Branches_Chart.xlsx")
PRONUNCIATION_GUIDE = Path("source-data/pronunciation_guide.txt")
MP3_OUT             = Path("40-branches.mp3")
HTML_OUT            = Path("40-branches.html")

GROUP_VOICE         = "en-US-GuyNeural"   # edge-tts male voice for headers

SILENCE_PRE_MS      = 4_000   # silence before each slide
BELL_HZ             = 880
BELL_DUR_MS         = 2_000
BELL_FADE_MS        = 1_800
BELL_POST_MS        = 300     # brief gap after bell before first spoken cell

GONG_HZ             = 220     # lower frequency distinguishes group headers
GONG_DUR_MS         = 3_000
GONG_FADE_MS        = 2_600

GAP_CELLS_MS        = 350     # between spoken cells within a slide row
GAP_REPEAT_MS       = 600     # before Vedic Name+Quality repeat at end of row

# pauses within group-header narration
GAP_GRP_INTRO_MS    = 300     # after gong, before group name
GAP_AFTER_NAME_MS   = 300     # after group name
GAP_WITHIN_MBMR_MS  = 500     # between name and quality of each member
GAP_BETWEEN_MBMR_MS = 800     # after quality, before next member
GAP_GRP_OUTRO_MS    = 1_500   # trailing silence after all members spoken

# ── group definitions ─────────────────────────────────────────────────────────
#   start / end = 0-based indices into the 42-entry entries[] list
#   display     = override which member indices to show in the header banner
#   notes       = footnote text keyed by member index
GROUPS = [
    {
        "name":  "Vedas",
        "desc":  "The four primordial expressions of Total Natural Law",
        "start": 0, "end": 3,
    },
    {
        "name":  "Vedangas",
        "desc":  "The six limbs of the Veda",
        "start": 4, "end": 9,
    },
    {
        "name":  "Upangas",
        "desc":  "The six philosophical systems, sub-limbs of the Veda",
        "start": 10, "end": 15,
    },
    {
        "name":  "Upa-Vedas",
        "desc":  "Twelve auxiliary Vedas, applied and embodied knowledge",
        "start": 16, "end": 27,
    },
    {
        "name":  "Brahmanas",
        "desc":  "Commentaries, chronicles, and the field of memory",
        "start": 28, "end": 35,
        "display": [28, 29, 30, 31, 34, 35],      # Ramayan & Mahabharat shown as Itihas sub-entries
        "notes":   {31: "incl. Ramayan and Mahabharat"},
    },
    {
        "name":  "Pratishakhyas",
        "desc":  "Six phonetic treatises, the finest layer of expression",
        "start": 36, "end": 41,
    },
]
GROUP_START_MAP = {g["start"]: gi for gi, g in enumerate(GROUPS)}

# ── pronunciation guide ───────────────────────────────────────────────────────
def _load_pronunciations() -> dict[str, str]:
    if not PRONUNCIATION_GUIDE.exists():
        print(f"  (no pronunciation guide found at {PRONUNCIATION_GUIDE})")
        return {}
    subs: dict[str, str] = {}
    for line in PRONUNCIATION_GUIDE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            word, phon = line.split("=", 1)
            subs[word.strip()] = phon.strip()
    print(f"  Loaded {len(subs)} pronunciation entries from {PRONUNCIATION_GUIDE.name}")
    return subs

PRONUNCIATIONS = _load_pronunciations()
_PRON_KEYS = sorted(PRONUNCIATIONS.keys(), key=len, reverse=True)

def roman_to_words(text: str) -> str:
    """Convert Roman numerals (I-XVIII) to spoken words in anatomical context."""
    _rom = {
        'XVIII':18,'XVII':17,'XVI':16,'XV':15,'XIV':14,'XIII':13,
        'XII':12,'XI':11,'X':10,'IX':9,'VIII':8,'VII':7,'VI':6,
        'V':5,'IV':4,'III':3,'II':2,'I':1,
    }
    _w = {1:'one',2:'two',3:'three',4:'four',5:'five',6:'six',
          7:'seven',8:'eight',9:'nine',10:'ten',11:'eleven',12:'twelve',
          13:'thirteen',14:'fourteen',15:'fifteen',16:'sixteen',
          17:'seventeen',18:'eighteen'}
    rpat = '|'.join(sorted(_rom, key=len, reverse=True))
    def r2w(s): return _w.get(_rom.get(s.upper(), 0), s)

    # context-sensitive: "Laminae I-VI", "Layer V", "Cortical Layer III", etc.
    ctx = r'(?:(?:Cortical\s+)?[Ll]ayer|[Ll]amina[e]?|Rexed)'
    # ranges first (longer match wins over single)
    text = re.sub(
        r'(' + ctx + r')\s+(' + rpat + r')-(' + rpat + r')',
        lambda m: m.group(1) + ' ' + r2w(m.group(2)) + ' through ' + r2w(m.group(3)),
        text
    )
    # single with context
    text = re.sub(
        r'(' + ctx + r')\s+(' + rpat + r')(?!\w)',
        lambda m: m.group(1) + ' ' + r2w(m.group(2)),
        text
    )
    # parenthetical ranges: (I-VI)
    text = re.sub(
        r'\((' + rpat + r')-(' + rpat + r')\)',
        lambda m: '(' + r2w(m.group(1)) + ' through ' + r2w(m.group(2)) + ')',
        text
    )
    return text

def preprocess(text: str) -> str:
    """Convert Roman numerals, apply pronunciation substitutions, fix word-number hyphens."""
    text = roman_to_words(text)
    for word in _PRON_KEYS:
        text = text.replace(word, PRONUNCIATIONS[word])
    text = re.sub(r"(\w)-(\d)", r"\1, \2", text)   # "Shrauta-20" >> "Shrauta, 20"
    return text

# ── audio helpers ─────────────────────────────────────────────────────────────
def sil(ms: int) -> AudioSegment:
    return AudioSegment.silent(duration=ms)

def tts(text: str) -> AudioSegment:
    """gTTS (regular female voice) for slide content."""
    buf = io.BytesIO()
    gTTS(text=preprocess(text), lang="en", slow=False).write_to_fp(buf)
    buf.seek(0)
    return AudioSegment.from_mp3(buf)

async def _edge_async(text: str) -> AudioSegment:
    buf = io.BytesIO()
    async for chunk in edge_tts.Communicate(preprocess(text), GROUP_VOICE).stream():
        if chunk["type"] == "audio":
            buf.write(chunk["data"])
    buf.seek(0)
    return AudioSegment.from_mp3(buf)

def tts_group(text: str) -> AudioSegment:
    """edge-tts male voice -- used only for group-header narration."""
    return asyncio.run(_edge_async(text))

def esc(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

BELL: AudioSegment = (
    Sine(BELL_HZ)
    .to_audio_segment(duration=BELL_DUR_MS, volume=-16)
    .fade_in(15).fade_out(BELL_FADE_MS)
)
GONG: AudioSegment = (
    Sine(GONG_HZ)
    .to_audio_segment(duration=GONG_DUR_MS, volume=-14)
    .fade_in(20).fade_out(GONG_FADE_MS)
)

# ── load spreadsheet ──────────────────────────────────────────────────────────
if not XLSX.exists():
    sys.exit(f"ERROR: {XLSX} not found.")

wb  = openpyxl.load_workbook(XLSX)
ws  = wb.active
raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
headers     = [str(h).strip() if h else f"Col{i+1}" for i, h in enumerate(raw_headers)]

entries: list[list[str]] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v).strip() if v is not None else "" for v in row]
    if any(vals):
        entries.append(vals)

print(f"Loaded {len(entries)} entries from {XLSX.name}")
print(f"Columns: {headers}\n")

# ── build group-header audio block ───────────────────────────────────────────
def build_group_header(gi: int) -> AudioSegment:
    g           = GROUPS[gi]
    display_idx = g.get("display") or list(range(g["start"], g["end"] + 1))
    notes       = g.get("notes", {})

    print(f"  >> group header: {g['name']}", flush=True)
    audio  = sil(1_000)
    audio += GONG
    audio += sil(GAP_GRP_INTRO_MS)
    audio += tts_group(g["name"])
    audio += sil(GAP_AFTER_NAME_MS)

    for mi in display_idx:
        row     = entries[mi]
        name    = row[0] if len(row) > 0 else ""
        quality = row[1] if len(row) > 1 else ""
        note    = notes.get(mi, "")

        spoken_name = name + (f", {note}" if note else "")
        if spoken_name:
            audio += tts_group(spoken_name)
            audio += sil(GAP_WITHIN_MBMR_MS)
        if quality:
            audio += tts_group(quality)
            audio += sil(GAP_BETWEEN_MBMR_MS)

    audio += sil(GAP_GRP_OUTRO_MS)
    return audio

# ── build main audio track ────────────────────────────────────────────────────
track              = sil(0)
slide_times_s: list[float]                     = []
group_header_ranges: list[tuple[float,float,int]] = []

for idx, row in enumerate(entries):
    vedic   = row[0] if row else ""
    quality = row[1] if len(row) > 1 else ""

    # ── group header before first slide of each group ─────────────────────
    if idx in GROUP_START_MAP:
        gi         = GROUP_START_MAP[idx]
        hdr_start  = len(track) / 1000.0
        track     += build_group_header(gi)
        hdr_end    = len(track) / 1000.0
        group_header_ranges.append((hdr_start, hdr_end, gi))

    print(f"  [{idx+1:2d}/{len(entries)}] {vedic}", flush=True)

    # ── normal slide audio ────────────────────────────────────────────────
    track += sil(SILENCE_PRE_MS)
    slide_times_s.append(len(track) / 1000.0)
    track += BELL
    track += sil(BELL_POST_MS)   # brief gap after bell before first spoken cell

    first_cell = True
    for col_name, val in zip(headers, row):
        if not val:
            continue
        if not first_cell:
            track += sil(GAP_CELLS_MS)
        track += tts(f"{col_name}: {val}")
        first_cell = False

    # repeat Vedic Name + Quality at end of each slide
    if vedic:
        track += sil(GAP_REPEAT_MS)
        repeat_text = f"{vedic}. {quality}." if quality else vedic
        track += tts(repeat_text)

dur_s      = len(track) / 1000.0
mins, secs = divmod(dur_s, 60)
print(f"\nTotal audio: {int(mins)}m {secs:.0f}s")
print(f"Exporting {MP3_OUT.name}...", flush=True)
track.export(str(MP3_OUT), format="mp3", bitrate="128k")
print(f"Saved: {MP3_OUT}")

# ── build HTML slideshow ──────────────────────────────────────────────────────
slides_html: list[str] = []
for i, row in enumerate(entries):
    vedic   = esc(row[0]) if len(row) > 0 else ""
    quality = esc(row[1]) if len(row) > 1 else ""
    aspect  = esc(row[2]) if len(row) > 2 else ""
    # Medical Definition (col 4) is spoken but not shown on slide

    inner = f'<p class="vedic">{vedic}</p>'
    if quality:
        inner += f'<p class="quality">{quality}</p>'
    if aspect:
        inner += f'<p class="aspect">{aspect}</p>'
    slides_html.append(f'  <div class="slide" id="s{i}">{inner}</div>')

# Group display metadata for the JS overlay
groups_js = json.dumps([
    {
        "name":    g["name"],
        "desc":    g["desc"],
        "start":   g["start"],
        "end":     g["end"],
        "display": g.get("display", list(range(g["start"], g["end"] + 1))),
        "notes":   g.get("notes", {}),
    }
    for g in GROUPS
])

SLIDES_BLOCK         = "\n".join(slides_html)
TIMES_JSON           = json.dumps(slide_times_s)
GHR_JSON             = json.dumps(group_header_ranges)   # [[start_s, end_s, gIdx], ...]
N                    = len(entries)
MP3_NAME             = MP3_OUT.name
SLIDE_DATA_JSON      = json.dumps([[row[0], row[1] if len(row) > 1 else ""] for row in entries])

html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<title>40 Branches of the Veda</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
html, body {{
  height: 100%;
  overflow: hidden;
  background: #0d1117;
  font-family: Georgia, 'Times New Roman', serif;
}}
#wrap {{
  display: flex;
  flex-direction: column;
  height: 100vh;
  height: 100dvh;
}}
#stage {{
  flex: 1;
  display: flex;
  align-items: center;
  justify-content: center;
  padding: 20px 24px 8px;
  overflow: hidden;
  position: relative;
}}

/* ── history panels ── */
#history {{
  position: absolute;
  top: 10px; left: 12px;
  max-width: 200px;
  max-height: calc(100% - 20px);
  overflow-y: auto;
  scrollbar-width: none;
  pointer-events: none;
  text-align: left;
}}
#history::-webkit-scrollbar {{ display: none; }}
#history-right {{
  position: absolute;
  top: 10px; right: 12px;
  max-width: 200px;
  max-height: calc(100% - 20px);
  overflow-y: auto;
  scrollbar-width: none;
  pointer-events: none;
  text-align: left;
}}
#history-right::-webkit-scrollbar {{ display: none; }}
.h-entry {{ margin-bottom: 5px; }}
.h-name  {{ font-size: .82rem; color: #5a7090; line-height: 1.3; }}
.h-qual  {{ font-size: .74rem; color: #3d5268; font-style: italic; line-height: 1.3; padding-left: 5px; }}
.h-group-label {{
  font-size: .68rem; color: #4a6080; letter-spacing: .05em;
  text-transform: uppercase;
  margin-top: 8px; margin-bottom: 2px;
  border-top: 1px solid #1e2a3a; padding-top: 4px;
}}
.h-group-label:first-child {{ margin-top: 0; border-top: none; padding-top: 0; }}

/* ── slides ── */
.slide {{
  display: none; flex-direction: column;
  align-items: center; justify-content: center;
  text-align: center; gap: 16px;
  max-width: 900px; width: 100%;
  animation: fi .5s ease;
}}
.slide.active {{ display: flex; }}
@keyframes fi {{
  from {{ opacity: 0; transform: translateY(8px); }}
  to   {{ opacity: 1; transform: translateY(0); }}
}}
.vedic   {{ color: #f5e6c8; font-size: clamp(2rem,6vw,3.6rem); font-weight: bold; letter-spacing: .04em; line-height: 1.2; }}
.quality {{ color: #7eb8d4; font-size: clamp(1.15rem,3.4vw,2.1rem); font-style: italic; line-height: 1.35; }}
.aspect  {{ color: #7dc47d; font-size: clamp(.92rem,2.5vw,1.35rem); line-height: 1.5; max-width: 740px; }}
#idle    {{ color: #3a4760; font-size: clamp(1rem,2.5vw,1.25rem); letter-spacing: .03em; }}

/* ── group header overlay ── */
#grp-hdr {{
  position: absolute; inset: 0;
  display: flex; flex-direction: column;
  align-items: center; justify-content: center;
  gap: 10px;
  background: rgba(10,14,20,0.94);
  z-index: 50; opacity: 0; pointer-events: none;
  transition: opacity 0.5s;
}}
#grp-hdr.show {{ opacity: 1; }}
.grp-name {{ font-size: clamp(1.5rem,4.5vw,2.6rem); color: #f5e6c8; letter-spacing: .12em; text-transform: uppercase; }}
.grp-sep  {{ width: 50%; border: none; border-top: 1px solid #1e2a3a; margin: 2px 0; }}
.grp-members {{ display: flex; flex-direction: column; gap: 6px; width: 100%; max-width: 820px; padding: 0 20px; }}
.grp-row  {{ display: flex; gap: 8px; justify-content: center; }}
.grp-item {{
  flex: 1; max-width: 260px; padding: 5px 10px; border-radius: 6px;
  text-align: center; font-size: clamp(.6rem,1.5vw,.8rem); line-height: 1.3; border: 1px solid;
}}
.grp-item.rishi    {{ color:#7eb8d4; border-color:#2a4a60; background:rgba(126,184,212,.07); }}
.grp-item.devata   {{ color:#d478a0; border-color:#60243a; background:rgba(212,120,160,.07); }}
.grp-item.chhandas {{ color:#c8a84b; border-color:#604820; background:rgba(200,168,75,.07);  }}
.grp-item.source   {{ color:#f5e6c8; border-color:#3a4760; background:rgba(245,230,200,.07); }}
.grp-item-quality  {{ display:block; font-size:.85em; opacity:.72; font-style:italic; margin-top:2px; }}
.grp-note {{ display:block; font-size:.75em; opacity:.65; margin-top:3px; font-style:italic; }}

/* ── audio bar ── */
#bar {{
  display: flex; align-items: center; gap: 10px;
  background: rgba(10,14,20,.97);
  padding: 8px 14px;
  padding-bottom: max(8px, env(safe-area-inset-bottom));
  border-top: 1px solid #1e2a3a;
}}
#playBtn {{
  width:48px; height:48px; background:#1e3048; border:1px solid #2a3a50;
  color:#f5e6c8; font-size:1.4rem; border-radius:50%; cursor:pointer;
  flex-shrink:0; -webkit-tap-highlight-color:transparent; touch-action:manipulation;
}}
#playBtn:active {{ background:#2a4060; }}
#prog {{ flex:1; accent-color:#7eb8d4; cursor:pointer; touch-action:none; height:4px; }}
#time-disp {{ color:#44506a; font-size:.72rem; font-family:monospace; white-space:nowrap; flex-shrink:0; }}
#ctr {{ color:#44506a; font-size:.78rem; font-family:monospace; white-space:nowrap; min-width:56px; text-align:right; flex-shrink:0; }}
#zone-top, #zone-mid, #zone-bot {{
  position:absolute; left:25%; right:25%;
  display:flex; align-items:center; justify-content:center;
  pointer-events:none; opacity:0;
  color:rgba(245,230,200,.22); font-size:clamp(.9rem,2.5vw,1.2rem);
  letter-spacing:.06em; transition:opacity .4s;
}}
#zone-top {{ top:0; height:33.33%; }}
#zone-mid {{ top:33.33%; height:33.34%; }}
#zone-bot {{ top:66.67%; height:33.33%; }}
@media (max-width:600px) {{ #history, #history-right {{ display:none; }} }}
</style>
</head>
<body>

<!-- password gate (hidden if session already unlocked) -->
<div id="gate" style="position:fixed;top:0;left:0;width:100%;height:100%;background:#0d1117;display:none;align-items:center;justify-content:center;z-index:1000;">
  <div style="text-align:center;color:#f5e6c8;font-family:Georgia,serif;">
    <p style="font-size:1.5rem;margin-bottom:1.2rem;">Enter password to continue</p>
    <input id="pwd" type="password" style="padding:8px 14px;font-size:1.1rem;border:1px solid #3a4760;background:#1a2232;color:#f5e6c8;border-radius:4px;outline:none;" placeholder="Password" autofocus />
    <br><br>
    <button id="enterBtn" style="padding:8px 24px;font-size:1rem;background:#7eb8d4;color:#0d1117;border:none;border-radius:4px;cursor:pointer;">Enter</button>
    <p id="gate-err" style="display:none;color:#e07070;margin-top:1rem;">Incorrect password. Please try again.</p>
  </div>
</div>

<div id="wrap">
  <div id="stage">
    <div id="idle">&#9654;&#xfe0e; Tap centre to play</div>
    <div id="history"></div>
    <div id="history-right"></div>
    <div id="grp-hdr">
      <div class="grp-name"></div>
      <hr class="grp-sep">
      <div class="grp-members"></div>
    </div>
    <div id="zone-top">&#9650; Faster</div>
    <div id="zone-mid">&#9654; Play / Pause</div>
    <div id="zone-bot">&#9660; Slower</div>
{SLIDES_BLOCK}
  </div>
  <div id="bar">
    <audio id="aud" preload="metadata" src="{MP3_NAME}"></audio>
    <button id="playBtn" aria-label="Play / Pause">&#9654;</button>
    <input id="prog" type="range" min="0" step="0.1" value="0" aria-label="Seek">
    <span id="time-disp">0:00</span>
    <span id="ctr"></span>
  </div>
</div>

<script>
// ── constants (generated by Python) ──────────────────────────────────────────
const TIMES              = {TIMES_JSON};
const GROUP_HEADER_RANGES= {GHR_JSON};   // [[start_s, end_s, groupIdx], ...]
const GROUPS             = {groups_js};
const SLIDE_DATA         = {SLIDE_DATA_JSON};
const N                  = {N};

// ── RDC quality per slide (matches spreadsheet order) ────────────────────────
const RDC = [
  'source','rishi','devata','chhandas',          // 0-3  Vedas
  'rishi','devata','chhandas','chhandas','devata','rishi',  // 4-9  Vedangas
  'rishi','devata','chhandas','chhandas','devata','rishi',  // 10-15 Upangas
  'rishi','devata','chhandas','chhandas','devata','rishi',  // 16-21 Upa-Vedas 1
  'rishi','devata','chhandas','chhandas','devata','rishi',  // 22-27 Upa-Vedas 2
  'rishi','devata','chhandas','chhandas',          // 28-31 Brahmanas 1
  'chhandas','chhandas','devata','rishi',          // 32-35 Brahmanas 2
  'rishi','devata','chhandas','chhandas','devata','rishi',  // 36-41 Pratishakhyas
];
const RDC_HIST = {{ rishi:'#3a7090', devata:'#903060', chhandas:'#806025', source:'#4a5a70' }};

// ── password gate ─────────────────────────────────────────────────────────────
(function initGate() {{
  const stored = sessionStorage.getItem('veda_pw_ok');
  if (stored === 'yes') return;   // already unlocked this session
  document.getElementById('gate').style.display = 'flex';
}})();

function checkPwd() {{
  const btn = document.getElementById('enterBtn');
  const err = document.getElementById('gate-err');
  if (document.getElementById('pwd').value === 'vedic2026') {{
    sessionStorage.setItem('veda_pw_ok', 'yes');
    document.getElementById('gate').style.display = 'none';
  }} else {{
    err.style.display = '';
    document.getElementById('pwd').value = '';
    document.getElementById('pwd').focus();
  }}
}}
document.getElementById('pwd').addEventListener('keydown', e => {{ if (e.key==='Enter') checkPwd(); }});
document.getElementById('enterBtn').addEventListener('click', checkPwd);

// ── core elements ─────────────────────────────────────────────────────────────
const aud      = document.getElementById('aud');
const idle     = document.getElementById('idle');
const ctr      = document.getElementById('ctr');
const hist     = document.getElementById('history');
const histRight= document.getElementById('history-right');
const grpHdr   = document.getElementById('grp-hdr');
let cur        = -1;
let curGrp     = -1;

// ── history panel ─────────────────────────────────────────────────────────────
function getGroupIdx(idx) {{
  for (let g=0; g<GROUPS.length; g++)
    if (idx >= GROUPS[g].start && idx <= GROUPS[g].end) return g;
  return -1;
}}
function makeEntries(arr, startIdx) {{
  let html = '', lastG = -1;
  arr.forEach(([n, q], i) => {{
    const absIdx = startIdx + i;
    const g = getGroupIdx(absIdx);
    if (g !== lastG) {{
      lastG = g;
      if (g >= 0) html += '<div class="h-group-label">[ ' + GROUPS[g].name + ' ]</div>';
    }}
    const col = RDC_HIST[RDC[absIdx] || 'source'];
    html += '<div class="h-entry">'
      + '<div class="h-name" style="color:' + col + '">' + n + '</div>'
      + '<div class="h-qual">' + q + '</div></div>';
  }});
  return html;
}}

// ── group header overlay (visual only -- audio is in the MP3) ─────────────────
function showGroupHeaderVisual(gi) {{
  if (gi === curGrp) return;
  curGrp = gi;
  const g = GROUPS[gi];
  const rdcOrder = {{ source:0, rishi:1, devata:2, chhandas:3 }};
  const dispIdx  = g.display || (()=>{{ const a=[];for(let i=g.start;i<=g.end;i++)a.push(i);return a;}})();
  const notes    = g.notes || {{}};
  const members  = dispIdx.map(i => ({{
    idx:     i,
    name:    SLIDE_DATA[i]?SLIDE_DATA[i][0]:'',
    quality: SLIDE_DATA[i]?SLIDE_DATA[i][1]:'',
    rdc:     RDC[i]||'source'
  }}));
  const rows=[]; for(let i=0;i<members.length;i+=3) rows.push(members.slice(i,i+3).sort((a,b)=>rdcOrder[a.rdc]-rdcOrder[b.rdc]));
  grpHdr.querySelector('.grp-name').textContent = g.name;
  grpHdr.querySelector('.grp-members').innerHTML = rows.map(row =>
    '<div class="grp-row">' + row.map(m =>
      '<div class="grp-item ' + m.rdc + '">' + m.name +
      (m.quality ? '<span class="grp-item-quality">' + m.quality + '</span>' : '') +
      (notes[m.idx] ? '<span class="grp-note">' + notes[m.idx] + '</span>' : '') +
      '</div>'
    ).join('') + '</div>'
  ).join('');
  grpHdr.classList.add('show');
}}
function hideGroupHeader() {{
  if (curGrp === -1) return;
  curGrp = -1;
  grpHdr.classList.remove('show');
}}

// ── slide show / hide ─────────────────────────────────────────────────────────
function show(idx) {{
  if (idx === cur) return;
  if (cur >= 0) {{
    const prev = document.getElementById('s'+cur);
    if (prev) prev.classList.remove('active','rdc-rishi','rdc-devata','rdc-chhandas','rdc-source');
  }}
  if (idx < 0) {{
    idle.style.display=''; ctr.textContent='';
    hist.innerHTML=''; histRight.innerHTML='';
  }} else {{
    idle.style.display='none';
    const el  = document.getElementById('s'+idx);
    const rdc = RDC[idx]||'source';
    if (el) el.classList.add('active','rdc-'+rdc);
    ctr.textContent      = (idx+1)+' / '+N;
    hist.innerHTML       = makeEntries(SLIDE_DATA.slice(0, Math.min(idx, 20)), 0);
    hist.scrollTop       = hist.scrollHeight;
    histRight.innerHTML  = idx > 20 ? makeEntries(SLIDE_DATA.slice(20, idx), 20) : '';
    histRight.scrollTop  = histRight.scrollHeight;
  }}
  cur = idx;
}}

// ── group-header overlay driven by audio position ────────────────────────────
function activeHeader(t) {{
  for (const [s,e,g] of GROUP_HEADER_RANGES)
    if (t >= s && t < e) return g;
  return -1;
}}

function sync() {{
  const t   = aud.currentTime;
  const hdr = activeHeader(t);
  if (hdr >= 0) {{
    showGroupHeaderVisual(hdr);
  }} else {{
    hideGroupHeader();
    let next = -1;
    for (let i=0; i<TIMES.length; i++) if (t >= TIMES[i]) next = i;
    show(next);
  }}
}}
function forceSync() {{
  document.querySelectorAll('.slide.active').forEach(el =>
    el.classList.remove('active','rdc-rishi','rdc-devata','rdc-chhandas','rdc-source'));
  cur = -1; curGrp = -1;
  grpHdr.classList.remove('show');
  idle.style.display = '';
  sync();
}}
aud.addEventListener('timeupdate', sync);
aud.addEventListener('seeked',     forceSync);
aud.addEventListener('ended',      () => {{ hideGroupHeader(); show(-1); }});

// ── screen wake lock ──────────────────────────────────────────────────────────
let wakeLock = null;
async function acquireWakeLock() {{
  if (!('wakeLock' in navigator)) return;
  try {{ wakeLock = await navigator.wakeLock.request('screen'); }} catch(e) {{}}
}}
function releaseWakeLock() {{
  if (wakeLock) {{ wakeLock.release().catch(()=>{{}}); wakeLock = null; }}
}}
aud.addEventListener('play',  acquireWakeLock);
aud.addEventListener('pause', releaseWakeLock);
aud.addEventListener('ended', releaseWakeLock);
document.addEventListener('visibilitychange', () => {{
  if (document.visibilityState === 'visible' && !aud.paused) acquireWakeLock();
}});

// ── audio interruption detection (timer-based for Google Maps etc.) ───────────
let userPaused    = false;
let interruptedAt = null;
let _resumeTimer  = null;

aud.addEventListener('pause', () => {{
  if (userPaused) return;           // user-initiated pause -- ignore
  interruptedAt = aud.currentTime;
  clearTimeout(_resumeTimer);
  _resumeTimer = setTimeout(() => {{
    if (aud.paused && !userPaused && interruptedAt !== null) {{
      aud.currentTime = Math.max(0, interruptedAt - 2);
      interruptedAt = null;
      aud.play().catch(()=>{{}});
    }}
  }}, 3500);
}});
aud.addEventListener('play', () => {{ interruptedAt = null; clearTimeout(_resumeTimer); }});
document.addEventListener('visibilitychange', () => {{
  if (document.visibilityState === 'visible' && interruptedAt !== null && !userPaused && aud.paused) {{
    clearTimeout(_resumeTimer);
    aud.currentTime = Math.max(0, interruptedAt - 2);
    interruptedAt = null;
    aud.play().catch(()=>{{}});
  }}
}});

// ── play/pause with user-intent tracking ─────────────────────────────────────
function togglePlayPause() {{
  if (aud.paused) {{ userPaused = false; interruptedAt = null; aud.play().catch(()=>{{}}); }}
  else            {{ userPaused = true;  aud.pause(); }}
}}

// ── playback speed ────────────────────────────────────────────────────────────
let speed = 1.0;
const MIN_SPEED = 0.25, MAX_SPEED = 3.0;
function setSpeed(s) {{
  speed = Math.min(MAX_SPEED, Math.max(MIN_SPEED, Math.round(s*100)/100));
  aud.playbackRate = speed;
  showBadge(speed.toFixed(2).replace(/\\.?0+$/,'')+'x');
}}

// ── UI helpers ────────────────────────────────────────────────────────────────
function showBadge(text) {{
  let b = document.getElementById('_badge');
  if (!b) {{
    b = document.createElement('div'); b.id='_badge';
    b.style.cssText='position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);'
      +'background:rgba(0,0,0,.72);color:#f5e6c8;font-size:1.6rem;padding:12px 28px;'
      +'border-radius:10px;pointer-events:none;z-index:600;font-family:Georgia,serif;'
      +'transition:opacity .45s;';
    document.body.appendChild(b);
  }}
  b.textContent=text; b.style.opacity='1';
  clearTimeout(b._t); b._t=setTimeout(()=>b.style.opacity='0',1100);
}}
function showSeekFlash(side) {{
  const el=document.createElement('div');
  el.textContent=side==='left'?'< 10s':'30s >';
  el.style.cssText='position:fixed;top:50%;'+(side==='left'?'left:12%':'right:12%')+';transform:translateY(-50%);'
    +'background:rgba(0,0,0,.65);color:#f5e6c8;font-size:1.3rem;padding:10px 22px;'
    +'border-radius:36px;pointer-events:none;opacity:1;transition:opacity .5s;z-index:600;font-family:Georgia,serif;';
  document.body.appendChild(el);
  setTimeout(()=>{{el.style.opacity='0';setTimeout(()=>el.remove(),520);}},680);
}}
function flashZone(el) {{
  el.style.opacity='1'; clearTimeout(el._zt);
  el._zt=setTimeout(()=>{{el.style.opacity='0';}},700);
}}

// ── bottom bar ────────────────────────────────────────────────────────────────
const playBtn  = document.getElementById('playBtn');
const prog     = document.getElementById('prog');
const timeDisp = document.getElementById('time-disp');
function fmtTime(s) {{
  s=Math.round(s||0); return Math.floor(s/60)+':'+String(s%60).padStart(2,'0');
}}
function updateBar() {{
  playBtn.innerHTML = aud.paused ? '&#9654;' : '&#9646;&#9646;';
  if (aud.duration) {{ prog.max=aud.duration; prog.value=aud.currentTime; }}
  timeDisp.textContent = fmtTime(aud.currentTime)+' / '+fmtTime(aud.duration);
}}
playBtn.addEventListener('click', () => togglePlayPause());
prog.addEventListener('input', () => {{ aud.currentTime = parseFloat(prog.value); }});
aud.addEventListener('play',          updateBar);
aud.addEventListener('pause',         updateBar);
aud.addEventListener('ended',         updateBar);
aud.addEventListener('timeupdate',    updateBar);
aud.addEventListener('loadedmetadata',updateBar);

// ── keyboard shortcuts ────────────────────────────────────────────────────────
document.addEventListener('keydown', function(e) {{
  switch (e.key) {{
    case ' ':         e.preventDefault(); togglePlayPause();  break;
    case 'ArrowRight':e.preventDefault(); aud.currentTime=Math.min(aud.currentTime+30,aud.duration||0); showSeekFlash('right'); break;
    case 'ArrowLeft': e.preventDefault(); aud.currentTime=Math.max(aud.currentTime-10,0); showSeekFlash('left'); break;
    case 'f': case 'F': e.preventDefault(); setSpeed(speed+0.25); break;
    case 's': case 'S': e.preventDefault(); setSpeed(speed-0.25); break;
  }}
}});

// ── tap zones ─────────────────────────────────────────────────────────────────
const zoneTop=document.getElementById('zone-top');
const zoneMid=document.getElementById('zone-mid');
const zoneBot=document.getElementById('zone-bot');
document.getElementById('stage').addEventListener('click', function(e) {{
  if (window.innerWidth > window.innerHeight && !document.fullscreenElement)
    document.documentElement.requestFullscreen?.().catch(()=>{{}});
  const rect=this.getBoundingClientRect();
  const relX=(e.clientX-rect.left)/rect.width;
  const relY=(e.clientY-rect.top)/rect.height;
  if      (relX < 0.25) {{ aud.currentTime=Math.max(0,aud.currentTime-10); showSeekFlash('left'); }}
  else if (relX > 0.75) {{ aud.currentTime=Math.min(aud.duration||0,aud.currentTime+30); showSeekFlash('right'); }}
  else if (relY < 0.333) {{ setSpeed(speed+0.25); flashZone(zoneTop); }}
  else if (relY < 0.667) {{ togglePlayPause(); flashZone(zoneMid); }}
  else                   {{ setSpeed(speed-0.25); flashZone(zoneBot); }}
}});
</script>
</body>
</html>"""

HTML_OUT.write_text(html_content, encoding="utf-8")
print(f"Saved: {HTML_OUT}")
print("\nAll done!")
print(f"  Audio  : {MP3_OUT.resolve()}")
print(f"  Slides : {HTML_OUT.resolve()}")
